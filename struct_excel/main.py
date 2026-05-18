from struct_excel.database import init_db, model_to_db
from struct_excel.normalization import normalize_sheet
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from struct_excel.reader import read_raw_row
from struct_excel.transform import (
    to_course,
    to_enrollment,
    to_session,
    to_student,
    to_supervisor,
)
import logging
import argparse
from pathlib import Path


def parse_args():
    parser = argparse.ArgumentParser(
        description="Process Excel workbook into normalized outputs."
    )
    parser.add_argument(
        "--src",
        required=True,
        help="Path to source Excel file",
    )

    return parser.parse_args()


def main():
    dist_dir = Path("dist")
    dist_dir.mkdir(parents=True, exist_ok=True)

    DB_PATH = "sqlite:///./dist/test.db"
    ERR_XLSX = "./dist/err.xlsx"

    engine = init_db(DB_PATH)

    args = parse_args()

    logging.basicConfig(
        level=logging.DEBUG,
        format="%(asctime)s %(levelname)s %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    logger = logging.getLogger(__name__)
    wb = load_workbook(args.src)
    err_wb = Workbook()

    for sheet in wb:
        err_ws = err_wb.create_sheet(sheet.title)
        normalize_sheet(sheet, err_ws)

    err_wb.save(ERR_XLSX)

    for sheet in wb:
        try:
            raw_rows = read_raw_row(sheet)
        except ValueError as e:
            logger.error(str(e))
            raise

        supervisors = to_supervisor(raw_rows)
        courses = to_course(raw_rows)
        students = to_student(raw_rows, supervisors)
        sessions = to_session(raw_rows, courses)
        enrollments = to_enrollment(raw_rows, students, courses, sessions)

        model_to_db(engine, supervisors)
        model_to_db(engine, courses)
        model_to_db(engine, students)
        model_to_db(engine, sessions)
        model_to_db(engine, enrollments)


if __name__ == "__main__":
    main()
